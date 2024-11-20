from robocorp.tasks import task
import logging
from helper import *
import openpyxl
from datetime import datetime
logger = logging.getLogger(__name__)

@task
def save_order_info(email, size, color, quantity, product_list, order_num):
    try:
        date_now = datetime.now().date()
        wb = open_workbook('data/order_detail.xlsx')
        save_order(wb, date_now, email, size, color, quantity)
        save_purchasing_info(wb, product_list, quantity, order_num, date_now)
        save_inventory_info(wb, product_list, quantity)
        output_file = 'output/output_order_detail.xlsx'
        save_workbook(wb, output_file)
        logger.info("Order information saved successfully")
        return output_file
    except Exception as e:
        logger.error(f"Failed to save information: {str(e)}")

def open_workbook(filepath):
    try:
        return openpyxl.load_workbook(filepath)
    except FileNotFoundError as e:
        logger.error(f"Workbook not found: {str(e)}")
        raise

def save_order(wb, date, email, size, color, quantity):
    sale_ws = wb['Sales Order']
    row = find_next_empty_row(sale_ws, start_row=3)
    data = [date, '', email, f"{size}, {color}", quantity]
    write_row(sale_ws, row, data)
    logger.info("Order added to Sales Order worksheet.")

def save_purchasing_info(wb, product_list, quantity, order_number, date):
    purchage_ws = wb['Purchasing']
    for product in product_list:
        row = find_next_empty_row(purchage_ws, start_row=3)
        data = ['Magneto', product['name'], product['size'], product['color'], quantity, product['status'], product['price'], date, order_number]
        write_row(purchage_ws, row, data)
    logger.info("Products added as new entries in Purchasing worksheet.")

def save_inventory_info(wb, product_list, quantity):
    inventory_ws = wb['Inventory']
    for product in product_list:
        row = find_row_by_product(inventory_ws, product['name'], 1)
        if row:
            current_quantity = inventory_ws.cell(row=row, column=2).value
            current_quantity = int(current_quantity) if current_quantity else 0
            new_quantity = current_quantity + int(quantity) 
            inventory_ws.cell(row=row, column=2, value=new_quantity)
        else:
            row = find_next_empty_row(inventory_ws, start_row=3)
            data = [product['name'], quantity]
            write_row(inventory_ws, row, data)
    logger.info("Products updated/added in Inventory worksheet.")

def find_next_empty_row(ws, start_row=3):
    """Find the next empty row in a worksheet starting from the given row."""
    while ws.cell(row=start_row, column=1).value:
        start_row += 1
    return start_row

def find_row_by_product(ws, product_name, item_column, start_row=3):
    """Finds a row with a specific product name. Returns None if not found."""
    for row in range(start_row, ws.max_row + 1):
        if ws.cell(row=row, column=item_column).value == product_name:
            return row
    return None

def write_row(ws, row, data):
    """Writes data to a specified row in a worksheet."""
    for col, value in enumerate(data, start=1):
        ws.cell(row=row, column=col, value=value)

def save_workbook(wb, filepath):
    wb.save(filepath)
    logger.info(f"Workbook saved to {filepath}")
