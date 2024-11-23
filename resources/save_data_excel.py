from robocorp.tasks import task
import logging
from helper import *
import openpyxl
from datetime import datetime
import time
logger = logging.getLogger(__name__)

@task
def save_order_info(email, quantity, product_list, order_num):
    try:
        date_now = datetime.now().date()
        wb = open_workbook('data/order_detail.xlsx')
        save_purchasing_info(wb, product_list, quantity, order_num, date_now)
        save_inventory_info(wb, product_list, quantity)
        save_order(wb, date_now, email, order_num)
        output_file = 'output/output_order_detail.xlsx'
        save_workbook(wb, output_file)
        logger.info("Order information saved successfully")
        return output_file
    except Exception as e:
        logger.error(f"Failed to save information: {str(e)}")

def open_workbook(filepath):
    try:
        return openpyxl.load_workbook(filepath)
    except FileNotFoundError as e:
        logger.error(f"Workbook not found: {str(e)}")
        raise

def save_order(wb, date, email, order_number):
    sale_ws = wb['Sales Order']
    row = find_order_row_by_date(sale_ws, date, start_row=3)
    sale_ws.cell(row=row, column=3, value=email)
    sale_ws.cell(row=row, column=10, value='Bought')
    sale_ws.cell(row=row, column=11, value=order_number)
    logger.info("Order added to Sales Order worksheet.")

def save_purchasing_info(wb, product_list, quantity, order_number, date):
    purchage_ws = wb['Purchasing']
    for product in product_list:
        row = find_next_empty_row(purchage_ws, start_row=3)
        data = ['Magneto', product['name'], product['size'], product['color'], quantity, product['status'], product['price'], date, order_number]
        write_row(purchage_ws, row, data)
    logger.info("Products added as new entries in Purchasing worksheet.")

def save_inventory_info(wb, product_list, quantity):
    inventory_ws = wb['Inventory']
    valid_product_list = [d for d in product_list if d['status'] == 'Added to cart']
    for product in valid_product_list:
        row = find_row_by_product(inventory_ws, product['name'], 1)
        if row:
            current_quantity = inventory_ws.cell(row=row, column=2).value
            current_quantity = int(current_quantity) if current_quantity else 0
            new_quantity = current_quantity + int(quantity) 
            inventory_ws.cell(row=row, column=2, value=new_quantity)
        else:
            row = find_next_empty_row(inventory_ws, start_row=3)
            data = [product['name'], quantity]
            write_row(inventory_ws, row, data)
    logger.info("Products updated/added in Inventory worksheet.")

def find_next_empty_row(ws, start_row=3):
    """Find the next empty row in a worksheet starting from the given row."""
    while ws.cell(row=start_row, column=1).value:
        start_row += 1
    return start_row

def find_row_by_product(ws, product_name, item_column, start_row=3):
    """Finds a row with a specific product name. Returns None if not found."""
    for row in range(start_row, ws.max_row + 1):
        if ws.cell(row=row, column=item_column).value == product_name:
            return row
    return None

def write_row(ws, row, data):
    """Writes data to a specified row in a worksheet."""
    for col, value in enumerate(data, start=1):
        ws.cell(row=row, column=col, value=value)

def save_workbook(wb, filepath):
    wb.save(filepath)
    logger.info(f"Workbook saved to {filepath}")

def get_sale_order_info(file_input):
    """
    Retrieve order information for the current date.
    """
    try:
        wb = open_workbook(file_input)
        sale_ws = wb['Sales Order']
        date_now = datetime.now().date()

        row = find_order_row_by_date(sale_ws, date_now)
        if not row:
            raise ValueError("No order found for today's date.")
        order_data, missing_fields = extract_order_data(sale_ws, row)
        if missing_fields:
            raise ValueError(f"Missing fields: {', '.join(missing_fields)}")
        logger.info(f"Order details: {order_data}")
        wb.close()
        return order_data

    except Exception as e:
        logger.error(f"Failed to retrieve sales order info: {str(e)}")
        raise


def find_order_row_by_date(ws, target_date, start_row=3):
    """
    Find row where the date matches the target date.
    """
    for row in range(start_row, ws.max_row + 1):
        cell_date = ws.cell(row=row, column=1).value
        if isinstance(cell_date, datetime):
            cell_date = cell_date.date()
        if cell_date == target_date:
            return row 
    return None


def extract_order_data(ws, row):
    """
    Extract order data from a specific row and check for missing fields.
    """
    fields = {
        'Category' : ws.cell(row=row, column=4).value,
        'Size': ws.cell(row=row, column=5).value,
        'Color': ws.cell(row=row, column=6).value,
        'Min Price': ws.cell(row=row, column=7).value,
        'Max Price': ws.cell(row=row, column=8).value,
        'Quantity': ws.cell(row=row, column=9).value,
    }
    missing_fields = [key for key, value in fields.items() if not value]
    order_data = {'Date': ws.cell(row=row, column=1).value}
    order_data.update(fields)

    return order_data, missing_fields


